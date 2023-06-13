from tkinter import  *
from pygame import *
import csv
from tkinter import messagebox # Para mostrar mensajes
from tkinter import ttk # Para usar los combobox
from tkinter import filedialog # Para abrir archivos
from PIL import Image, ImageTk # Para usar imagenes
import PIL.Image # Ajusta el tamaño de la imagen
import os # Para abrir archivos
import time # Para usar el reloj
import datetime # Para usar el reloj
import threading # Para usar el reloj
import openpyxl # Para usar archivos de excel
from openpyxl.chart import  Reference # Para usar archivos de excel
from matplotlib import pyplot as plt # Para las graficas
from matplotlib import style # Para las graficas
from matplotlib import animation # Para las graficas
from matplotlib import get_backend


lista = [] # Lista que contendra los datos de los empleados
lista2 = [] #donde se guardan los datos de los empleados que se van a eliminar o modificar
ultima = [] # Ultima vez de ejecucion del codigo

"""
Imagenes: esta funcion pondra imagenes con un tamaño determinado
E: una imagen
R: png o gif
S: la imagen con el tamaño indicado
"""
def Imagenes(img,size):
    ruta = None
    if size != None:
        ruta = PIL.Image.open("Adds/"+img).resize((size))
    else:
        ruta = PIL.Image.open("Adds/"+img)
    imagen = ImageTk.PhotoImage(ruta)
    return imagen 
    
"""
leer:Funcion que lee el archivo de texto con las caracteristica del robot
E: Nada
R: Nada
S: Una lista con las caracteristicas de los empleados
"""
def leer():
    with open("Datos.csv",'r') as csvfile:
        csvreader = csv.reader(csvfile)
        for line in csvreader:
            lista.append(line)

"""
Funcion que escribe en el archivo de texto(Actualiza el archivo)
E: Nada
R: Nada
S: Nada
"""
def escribir():
    with open("Datos.csv",'w',newline="") as csvfile:
        csvwriter = csv.writer(csvfile)
        csvwriter.writerows(lista)

"""
leer2:Funcion que lee el archivo de texto .txt con la ultima vez de ejecucion del codigo
E: Nada
R: Nada
S: Una lista con la ultima vez de ejecucion del codigo
"""
def leer2():
    with open("Ultima.txt",'r') as txtfile:
        for line in txtfile:
            ultima.append(line)
"""
actualizar:Funcion que actualiza el archivo de texto .txt con la ultima vez de ejecucion del codigo
E: Nada
R: Nada
S: Nada
"""
def actualizar():
    with open("Ultima.txt",'w',newline="") as txtfile:
        txtfile.write(str(datetime.datetime.now().strftime("%d/%m/%Y")))

"""
Salario_Total: Funcion que calcula el salario total de un empleado
E: la lista de empleados
R: Numeros
S: El salario total
"""
def Salario_total():
    for i in lista:
        if i[7] == "0":
            i[9] = str((float(i[3]) * float(i[4]) - (float(i[3]) * float(i[4])) * 0.15) * 2)
        else:
            i[9] = str((float(i[3]) * float(i[4]) + float(i[3]) * float(i[4]) * 0.15) * float(i[7]) * 2)
            # reedondear a 2 decimales
            i[9] = str(round(float(i[9]),2))
    escribir()

""""
Excel: Funcion que crea un archivo de excel con los datos de los empleados y se actualiza cada vez que se ejecuta
 el codigo
E: Nada
R: Nada
S: Nada
"""	
def Excel():
    wb = openpyxl.load_workbook("DatosE.xlsx")
    ws = wb["Datos Empleados"]
    ws.delete_rows(1,ws.max_row)

    ws.append(["Nombre","Apellido","Codigo","Sueldo","Horas","Sexo","Edad","Horas Extra","Fecha de Contratacion",
    "Salario Total"])

    # Se ajusta el tamaño de las columnas y su fuente y se centrar el texto
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 20

    ws['A1'].font = openpyxl.styles.Font(size=14)
    ws['B1'].font = openpyxl.styles.Font(size=14)
    ws['C1'].font = openpyxl.styles.Font(size=14)
    ws['D1'].font = openpyxl.styles.Font(size=14)
    ws['E1'].font = openpyxl.styles.Font(size=14)
    ws['F1'].font = openpyxl.styles.Font(size=14)
    ws['G1'].font = openpyxl.styles.Font(size=14)
    ws['H1'].font = openpyxl.styles.Font(size=14)
    ws['I1'].font = openpyxl.styles.Font(size=14)
    ws['J1'].font = openpyxl.styles.Font(size=14)

    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['B1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['C1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['D1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['E1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['F1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['G1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['H1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['I1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    ws['J1'].alignment = openpyxl.styles.Alignment(horizontal="center")

    # Se agregan los datos de los empleados y se cetran en las celdas el resto de los datos
    for i in lista:
        ws.append(i)
    for i in range(2,len(lista)+2):
        ws['A'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws['B'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws['C'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws['D'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws['E'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws['F'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws['G'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws['H'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws['I'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")
        ws['J'+str(i)].alignment = openpyxl.styles.Alignment(horizontal="center")

    wb.save("DatosE.xlsx") 

"""
Resetear_H_E: Funcion que resetea las horas extra de los empleados, si se cumple algunas de las siguientes condiciones
1. Si se esta en el dia 15 o 30 del mes
2. Si la ultima vez que se resetearon las horas extra fue hace mas de 15 dias
E: Nada
R: Nada
S: Nada
"""
def Resetear_H_E():
    if  str(datetime.datetime.now().strftime("%d")) == "15" or str(datetime.datetime.now().strftime("%d")) == "30":
        for i in lista:
            if i[7] != "0":
                i[7] = "0"
        escribir()
    elif int(datetime.datetime.now().strftime("%m"))> int(ultima[0][3:5]):
        for i in lista:
            if i[7] != "0":
                i[7] = "0"
        escribir()
    
    elif int(datetime.datetime.now().strftime("%Y")) > int(ultima[0][6:10]):
        for i in lista:
            if i[7] != "0":
                i[7] = "0"
        escribir()

    elif int(datetime.datetime.now().strftime("%d")) - int(ultima[0][0:2]) > 15:
        for i in lista:
            if i[7] != "0":
                i[7] = "0"
        escribir()
    
    else:
        pass


##################################################################

"""
class Empresa(): ESta es una clase que simula una empresa, la cual tiene varios atributos y metodos
que se le asignan a cada empleado
"""

class Empresa():
    def __init__(self, nombre, apellido, codigo, sueldo, horas, sexo, edad, horas_extra,contratacion,total):
        self.nombre = nombre # Nombre del empleado
        self.apellido = apellido # Apellido del empleado
        self.codigo = codigo # Codigo del empleado
        self.sueldo = sueldo # Sueldo del empleado(base)
        self.horas = horas # Horas trabajadas del empleado por semana
        self.sexo = sexo # Sexo del empleado
        self.edad = edad # Edad del empleado
        self.contratacion = contratacion # Fecha de contratacion del empleado
        self.horas_extra = horas_extra # Horas extra trabajadas por el empleado
        self.total = total # Salario total del empleado
    
    """
    Aqui todos los metodos get, que son los que se encargan de retornar los atributos de la clase
    """

    def getNombre(self):
        return self.nombre

    def getApellido(self):
        return self.apellido
    
    def getCodigo(self):
        return self.codigo

    def getSueldo(self):
        return self.sueldo

    def getHoras(self):
        return self.horas

    def getSexo(self):
        return self.sexo
    
    def getEdad(self):
        return self.edad

    def getHorasExtra(self):
        return self.horas_extra

    def getContratacion(self):
        return self.contratacion

    def getTotal(self):
        return self.total

    """	
    Aqui todos los metodos set, que son los que se encargan de modificar los atributos de la clase
    """

    def setNombre(self, nombre):
        self.nombre = nombre
    
    def setApellido(self, apellido):
        self.apellido = apellido

    def setCodigo(self, codigo):
        self.codigo = codigo
    
    def setSueldo(self, sueldo):
        self.sueldo = sueldo
    
    def setHoras(self, horas):
        self.horas = horas
    
    def setSexo(self, sexo):
        self.sexo = sexo

    def setEdad(self, edad):
        self.edad = edad
    
    def setHorasExtra(self, horas_extra):
        self.horas_extra = horas_extra
    
    def setContratacion(self, contratacion):
        self.contratacion = contratacion

    def setTotal(self, total):
        self.total = total

    """
    __str__: Esta funcion se encarga de retornar los atributos de la clase en forma de string
    E: Nada
    R: Nada
    S: Un string con los atributos de la clase
    """
    def __str__(self):
        return ("Nombre: " + self.nombre + " Apellido: " 
                + self.apellido +" Codigo: " + str(self.codigo) 
                + " Sueldo: " + str(self.sueldo) + " Horas: " + str(self.horas) 
                + " Sexo: " + self.sexo + " Edad: " + str(self.edad) +" Horas Extra" + str(self.horas_extra)
                +" Contratacion: " + self.contratacion + " Total: " + str(self.total))

"""
class Ventana_Añadir_Empleado(): Esta clase es la que se encarga de crear la ventana Toplevel y los widgets que se van a 
mostrar en la ventana
Esta clase tiene varias funciones ademas de la __init__, las cuales son de interfaz o de validacion de datos, 
su funcion principal es la de añadir un empleado a la lista de empleados
"""
#Clase que crea la ventana para añadir un empleado
class Ventana_Añadir_Empleado():

    def __init__(self,ventana):

        self.Ventana = ventana.withdraw()
        self.ventana = Toplevel()
        self.ventana.title("Añadir Empleado")
        self.ventana.geometry("500x500+500+100")
        self.ventana.resizable(0,0)
        self.ventana.config(bg="Black")
        self.ventana.iconbitmap("Adds/icon.ico")	
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")

        # Variables
        self.nombre = StringVar()
        self.apellido = StringVar()
        self.codigo = StringVar()
        self.sueldo = StringVar()
        self.horas = StringVar()
        self.edad = StringVar()
        self.sexo = IntVar()
        self.mensaje = StringVar()
        self.contratacion = str(datetime.datetime.now().strftime("%d/%m/%Y"))
        self.horas_extra = "0"
        self.total = "0"


        self.crearWidgets()

    """
    def crearWidgets(): Esta funcion es la que se encarga de crear los widgets que se van a mostrar en la ventana
    """
    def crearWidgets(self):

        self.botonVolver = Button(self.ventana, text="Volver", font=("Arial", 15), command=self.volver)
        self.botonVolver.place(x=375, y=410)

        self.titulo = Label(self.ventana, text="Empresa", font=("Arial", 20), bg="blue", fg="white")
        self.titulo.place(x=200, y=10)

        self.nombreLabel = Label(self.ventana, text="Nombre: ", font=("Arial", 15), bg="blue", fg="white")
        self.nombreLabel.place(x=10, y=50)
        self.nombreEntry = Entry(self.ventana, textvariable=self.nombre, font=("Arial", 15))
        self.nombreEntry.place(x=100, y=50)

        self.apellidoLabel = Label(self.ventana, text="Apellido: ", font=("Arial", 15), bg="blue", fg="white")
        self.apellidoLabel.place(x=10, y=100)
        self.apellidoEntry = Entry(self.ventana, textvariable=self.apellido, font=("Arial", 15))
        self.apellidoEntry.place(x=100, y=100)

        self.codigoLabel = Label(self.ventana, text="Codigo: ", font=("Arial", 15), bg="blue", fg="white")
        self.codigoLabel.place(x=10, y=150)
        self.codigoEntry = Entry(self.ventana, textvariable=self.codigo, font=("Arial", 15))
        self.codigoEntry.place(x=100, y=150)

        self.sueldoLabel = Label(self.ventana, text="Sueldo: ", font=("Arial", 15), bg="blue", fg="white")
        self.sueldoLabel.place(x=10, y=200)
        self.sueldoEntry = Entry(self.ventana, textvariable=self.sueldo, font=("Arial", 15))
        self.sueldoEntry.place(x=100, y=200)

        self.horasLabel = Label(self.ventana, text="Horas/S: ", font=("Arial", 15), bg="blue", fg="white")
        self.horasLabel.place(x=10, y=250)
        self.horasEntry = Entry(self.ventana, textvariable=self.horas, font=("Arial", 15))
        self.horasEntry.place(x=100, y=250)

        self.edadLabel = Label(self.ventana, text="Edad: ", font=("Arial", 15), bg="blue", fg="white")
        self.edadLabel.place(x=10, y=300)
        self.edadEntry = Entry(self.ventana, textvariable=self.edad, font=("Arial", 15))
        self.edadEntry.place(x=100, y=300)

        self.sexoLabel = Label(self.ventana, text="Sexo: ", font=("Arial", 15), bg="blue", fg="white")
        self.sexoLabel.place(x=10, y=350)
        self.opcion_masculino = Radiobutton(self.ventana, text="Masculino", variable=self.sexo, value=1)
        self.opcion_femenino = Radiobutton(self.ventana, text="Femenino", variable=self.sexo, value=2)
        self.opcion_otro = Radiobutton(self.ventana, text="Otro", variable=self.sexo, value=3)

        self.opcion_masculino.place(x=100, y=350)
        self.opcion_femenino.place(x=200, y=350)
        self.opcion_otro.place(x=300, y=350)

        self.mensajeLabel = Label(self.ventana, textvariable=self.mensaje, font=("Arial", 15), bg="blue", fg="white")
        self.mensajeLabel.place(x=10, y=400)

        self.boton = Button(self.ventana, text="Enviar", font=("Arial", 15), command=self.enviar)
        self.boton.place(x=200, y=400)

    """
    Esta funcion se encarga de validar los datos que se ingresan en los Entry y los Radiobutton
    Tambien se encarga de enviar los datos para que estos sean guardados en el archivo
    E: los datos que se ingresan en los Entry y los Radiobutton
    S: Guardar los datos en el archivo
    R: Los datos no pueden estar vacios, ni pueden ser invalidos
    """
    def enviar(self):
        nombre = self.nombre.get()
        apellido = self.apellido.get()
        codigo = self.codigo.get()
        sueldo = self.sueldo.get()
        horas = self.horas.get()
        edad = self.edad.get()
        sexo = self.sexo.get()
        contratacion =  self.contratacion
        horas_extra = self.horas_extra
        total = self.total

        #Condicionales que se tienen que cumplir para que los datos sean guardados en el archivo

        if nombre == "" or  apellido == "" or codigo == "" or sueldo == "" or horas == ""  or edad == "" or sexo == 0:
            messagebox.showerror("Error", "Debe llenar todos los campos")
        else:
            
            if codigo != "":
                try:

                    codigo2 = int(codigo)
                    if len(codigo) != 4:
                        messagebox.showerror("Error", "El codigo debe tener 4 digitos")
                        self.codigo.set("")
                        return False
                    else:
                        for i in lista:
                            if i[1] == codigo:
                                messagebox.showerror("Error", "El codigo ya existe")
                                self.codigo.set("")
                                return False
                            else:
                                continue
 
                except ValueError:
                    messagebox.showerror("Error", "El codigo debe ser un numero")
                    self.codigo.set("")
                    return False
            
            if sueldo != "":
                try:
                    sueldo = float(sueldo)
                except:
                    messagebox.showerror("Error", "El sueldo debe ser un numero")
                    self.sueldo.set("")
                    return False
            
            if horas != "":
                try:
                    horas = float(horas)
                    if horas > 48:
                        messagebox.showerror("Error", "Las horas no pueden ser mayor a 48")
                        self.horas.set("")
                        return False
                    
                except ValueError:
                    messagebox.showerror("Error", "Las horas deben ser un numero")
                    self.horas.set("")
                    return False
            
            if edad != "":
                try:
                    edad = int(edad)
                    if edad < 18 or edad > 65:
                        messagebox.showerror("Error", "La edad debe estar entre 18 y 65")
                        self.edad.set("")
                        return False
                    elif edad > 65:
                        messagebox.showinfo("Info", "Por favor solicite una jubilacion")
                        self.edad.set("")
                        return False
                except ValueError:
                    messagebox.showerror("Error", "La edad debe ser un numero")
                    self.edad.set("")
                    return False
            
            if sexo == 1:
                sexo = "Masculino"
            elif sexo == 2:
                sexo = "Femenino"
            elif sexo == 3:
                sexo = "Otro"
            else:
                messagebox.showerror("Error", "Debe seleccionar un sexo")
                return False
            
            if nombre[0].isalpha() == False:
                messagebox.showerror("Error", "El nombre debe comenzar con una letra")
                self.nombre.set("")
                return False
            
            if apellido[0].isalpha() == False:
                messagebox.showerror("Error", "El apellido debe comenzar con una letra")
                self.apellido.set("")
                return False
            
            if nombre[0] != nombre[0].upper():
                messagebox.showerror("Error", "El nombre debe comenzar con mayuscula")
                self.nombre.set("")
                return False

            for i in nombre:
                if i == " ":
                    messagebox.showerror("Error", "El nombre no debe contener espacios")
                    self.nombre.set("")
                    return False
            
            if apellido[0] != apellido[0].upper():
                messagebox.showerror("Error", "El apellido debe comenzar con mayuscula")
                self.apellido.set("")
                return False
            
            if nombre.isalpha() == False:
                messagebox.showerror("Error", "El nombre no debe contener numeros ni caracteres especiales")
                self.nombre.set("")
                return False
            
            if apellido.isalpha() == False:
                messagebox.showerror("Error", "El apellido no debe contener numeros ni caracteres especiales")
                self.apellido.set("")
                return False

            for i in apellido:
                if i == " ":
                    messagebox.showerror("Error", "El apellido no debe contener espacios")
                    self.apellido.set("")
                    return False     

            self.mensaje.set("Datos enviados")
            self.nombre.set("")
            self.apellido.set("")
            self.codigo.set("")
            self.sueldo.set("")
            self.horas.set("")
            self.sexo.set("")
            self.edad.set("")
            
            
            self.empresa = Empresa(nombre, apellido, codigo, sueldo, horas, sexo, edad,
                horas_extra,contratacion,total)

            print(self.empresa)
            
            lista.append([self.empresa.nombre, self.empresa.apellido,self.empresa.codigo,
             self.empresa.sueldo, self.empresa.horas, self.empresa.sexo, self.empresa.edad,
             self.empresa.horas_extra,self.empresa.contratacion,self.empresa.total])

            Salario_total()
            escribir()
            return True
    """
    def volver(self): Esta funcion cierra la ventana actual y abre la ventana principal
    """
    def volver(self):
        self.ventana.destroy()
        self.ventana = Aplicacion()

"""
class Ventana_Jefe(): Esta clase abre una ventana donde estaran los botones para abrir las otras ventanas que solo puede
abrir el administrador o jefe de la empresa
"""	

#Clase para la ventana del jefe
class Ventana_Jefe():

    def __init__(self,ventana):

        self.Ventana = ventana.withdraw()
        self.ventana = Toplevel()
        self.ventana.geometry("500x500+500+100")
        self.ventana.title("Jefe")
        self.ventana.resizable(0,0)
        self.ventana.config(bg="Black")
        self.ventana.iconbitmap("Adds/icon.ico")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")

        #Creacion de los botones, labels y entrys

        self.boton1 = Button(self.ventana, text="Añadir Usuario", font=("Arial", 15), command=self.añadir_usuario)
        self.boton1.place(x=200, y=100)

        self.boton2 = Button(self.ventana, text="Calcular Salario", font=("Arial", 15), command=self.Verificar_Monto)
        self.boton2.place(x=200, y=200)

        self.boton3 = Button(self.ventana, text="Ordenar por", font=("Arial", 15), command=self.Ordenar)
        self.boton3.place(x=200, y=300)

        self.boton4 = Button(self.ventana, text="Salir", font=("Arial", 15), command=self.ventana.destroy)
        self.boton4.place(x=200, y=500)

        self.boton5 = Button(self.ventana, text="Retirar Empleado", font=("Arial", 15), command=self.Retirar)
        self.boton5.place(x=200, y=400)

        self.boton6 = Button(self.ventana, text="Volver", font=("Arial", 15), command=self.Volver)
        self.boton6.place(x=0, y=0)

    """
    def añadir_usuario(self): Esta funcion abre una ventana donde se podra añadir un usuario la cual es la class Ventana
    que se hizo anterormente y cierra la ventana principal
    """	
    def añadir_usuario(self):
        self.ventana1 = Ventana_Añadir_Empleado(self.ventana)

    """
    def Verificar_Monto(self): Esta funcion abre una ventana donde se podra calcular el salario de los empleados
    """
    def Verificar_Monto(self):
        self.ventana2 = C_Salario(self.ventana)
    
    """
    def Ordenar(self): Esta funcion abre una ventana donde se podra ordenar los 
    empleados por nombre, apellido, codigo, sueldo, horas, sexo y edad
    """
    def Ordenar(self):
        self.ventana3 = Ventana_Ordenar()
    
    """
    def Retirar(self): Esta funcion abre una ventana donde se podra retirar un empleado y calcular su salario final
    """
    def Retirar(self):
        self.ventana4 = Ventana_Retirar()

    """	
    def Volver(self): Esta funcion abre una ventana inicial
    """	
    def Volver(self):
        self.ventana.destroy()
        self.ventana = Aplicacion()

"""
class Ventana_Empleado(): Esta clase abre una ventana donde estaran los botones para abrir las otras ventanas disponibles 
para los empleados
"""
# Esta ventana solo la puede abrir un empleado
class Ventana_Empleado():

    def __init__(self,ventana):

        self.Ventana = ventana.withdraw()
        self.ventana = Toplevel()
        self.ventana.geometry("500x500+500+100")
        self.ventana.title("Ventana Principal")
        self.ventana.resizable(0,0)
        self.ventana.config(bg="Black")
        self.ventana.iconbitmap("Adds/icon.ico")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")

        #Creacion de los botones, labels y entrys

        self.boton1 = Button(self.ventana, text="Calcular Salario", font=("Arial", 15), command=self.Verificar_Monto)
        self.boton1.place(x=150, y=100)

        self.boton3 = Button(self.ventana, text="Añadir Empleado", font=("Arial", 15), command=self.añadir_usuario)
        self.boton3.place(x=150, y=200)

        self.boton4 = Button(self.ventana, text="Volver", font=("Arial", 15), command=self.volver)
        self.boton4.place(x=0, y=0)

    """
    def Verificar_Monto(self): Esta funcion abre una ventana donde se podra calcular el salario del empleado que esta
    logeado y otras funciones
    """
    def Verificar_Monto(self):
        self.ventana1 = Salario_Empleado(self.ventana)
    
    """
    def añadir_usuario(self): Esta funcion abre una ventana donde se podra añadir un usuario 
    """
    def añadir_usuario(self):
        self.ventana2 = Ventana_Añadir_Empleado(self.ventana)

    """
    def volver(self): Esta funcion abre una la ventana principal
    """
    def volver(self):
        self.ventana.destroy()
        self.Ventana = Aplicacion()

"""
Clase C_Salario(): Esta clase es la que se encarga de calcular el salario de los empleados y tambien se pede solicitar
el de un empleado en especifico, esta clase tomara de la lista de la clase Ventana para calcular el salario y mostrar las
respectivas tablas
"""
class C_Salario():

    def __init__(self,ventana):

        self.Ventana = ventana.withdraw()
        self.ventana = Toplevel()
        self.ventana.geometry("500x500+500+100")
        self.ventana.title("Calcular Salario")
        self.ventana.resizable(0,0)
        self.ventana.config(bg="Black")
        self.ventana.iconbitmap("Adds/icon.ico")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        
        self.codigo = StringVar()
        self.mensaje = StringVar()
        self.mensaje.set("")

        #Creacion de los botones, labels y entrys

        self.label1 = Label(self.ventana, text="Codigo", font=("Arial", 15), bg="blue", fg="white")
        self.label1.place(x=50, y=50)

        self.entrada1 = Entry(self.ventana, textvariable=self.codigo, font=("Arial", 15))
        self.entrada1.place(x=150, y=50)

        self.boton1 = Button(self.ventana, text="Salario Especifico", font=("Arial", 15), command=self.calcular)
        self.boton1.place(x=175, y=100)

        self.boton2 = Button(self.ventana, text="Calcular Todos", font=("Arial", 15), command=self.calcular_todo)
        self.boton2.place(x=190, y=150)

        self.boton3 = Button(self.ventana, text="Horas Extra", font=("Arial", 15), command=self.horas_extra)
        self.boton3.place(x=190, y=200)

        self.boton4 = Button(self.ventana, text="Volver", font=("Arial", 15), command=self.volver)
        self.boton4.place(x=0, y=0)

        self.boton5 = Button(self.ventana, text="Graficas", font=("Arial", 15), command=self.graficas)
        self.boton5.place(x=200, y=250)

        self.label2 = Label(self.ventana, textvariable=self.mensaje, font=("Arial", 15), bg="blue", fg="white")
        self.label2.place(x=50, y=350)
    
    ############################## FUNCIONES INDEPENDIENTES A LAS HORAS EXTRA ##############################
    """
    def calcular(self): Esta funcion se encarga de calcular el salario de un empleado en especifico
    E: codigo
    S: el salario del empleado
    R: el codigo debe existir
    """	
    def calcular(self):
        codigo = self.codigo.get()
        if codigo == "":
            self.mensaje.set("El codigo no puede estar vacio")
            return False
        else:
            for i in lista:
                if codigo == i[2]:
                    messagebox.showinfo("Salario", "El salario de " + i[0] + i[1] +
                    " es de " +  str(i[9]))
                    return True
            messagebox.showerror("Error", "El codigo no existe")
            return False
                
    """
    def calcular_todo(self): Esta funcion se encarga de calcular el salario de todos los empleados
    E: lista de empleados
    S: el salario de todos los empleados
    R: la lista no puede estar vacia
    """	
    def calcular_todo(self):
        self.ventana2 = Toplevel()  
        self.ventana2.geometry("1350x700+0+0")
        self.ventana2.title("Calcular Salario")
        self.ventana2.resizable(0,0)
        self.ventana2.config(bg="blue")
        self.ventana.iconbitmap("Adds/icon.ico")
        self.ventana2.config(bd=25)
        self.ventana2.config(relief="groove")
        self.ventana2.config(cursor="pirate")
        self.ventana2.config(bd=25)
        self.ventana2.config(relief="groove")
        self.ventana2.config(cursor="pirate")

        self.tabla = ttk.Treeview(self.ventana2, height=30, columns=("col1", "col2", "col3", "col4", "col5", "col6"
        , "col7", "col8", "col9", "col10", "col11"))
        self.tabla.column("#0", width=100,anchor=CENTER)
        self.tabla.column("col1", width=150,anchor=CENTER)
        self.tabla.column("col2", width=150,anchor=CENTER)
        self.tabla.column("col3", width=100,anchor=CENTER)
        self.tabla.column("col4", width=100,anchor=CENTER)
        self.tabla.column("col5", width=100,anchor=CENTER)
        self.tabla.column("col6", width=100,anchor=CENTER)
        self.tabla.column("col7", width=100,anchor=CENTER)
        self.tabla.column("col8", width=100,anchor=CENTER)
        self.tabla.column("col9", width=100,anchor=CENTER)
        self.tabla.column("col10", width=100,anchor=CENTER)
        self.tabla.column("col11", width=100,anchor=CENTER)
        
        self.tabla.heading("#0", text="Empleado", anchor=CENTER)
        self.tabla.heading("col1", text="Apellido", anchor=CENTER)
        self.tabla.heading("col2", text="Codigo", anchor=CENTER)
        self.tabla.heading("col3", text="Sueldo por hora", anchor=CENTER)
        self.tabla.heading("col4", text="Horas por Semana", anchor=CENTER)
        self.tabla.heading("col5", text="Sexo", anchor=CENTER)
        self.tabla.heading("col6", text="Edad", anchor=CENTER)
        self.tabla.heading("col7", text="Reduccion $", anchor=CENTER)
        self.tabla.heading("col8", text="Sub total $", anchor=CENTER)
        self.tabla.heading("col9", text="Horas Extra", anchor=CENTER)
        self.tabla.heading("col10", text="Dia de cotrato", anchor=CENTER)
        self.tabla.heading("col11", text="Total", anchor=CENTER)

        self.tabla.place(x=0, y=0)
        self.tabla.pack()

        # Se encarga de mostrar los datos en la tabla, los datos de la clomna 7 y 8 solo tienen 2 decimales
        for i in lista:
            self.tabla.insert("", 0, text=i[0], values=(i[1], i[2],  i[3], i[4], i[5], i[6], 
             str(round(float(i[3]) * float(i[4]) * 0.15, 2)) ,
             str(round(float(i[3]) * float(i[4]) - float(i[3]) * float(i[4]) * 0.15,2)),i[7], i[8], i[9]))

    ############################## FUNCIONES DE HORAS EXTRA ##############################
    """
    def horas_extra(self): Esta funcion se encarga de sumar las horas extra de los empleados ya sea de uno en especifico
    o de todos o de solo algunos empleados
    """
    def horas_extra(self):
        self.ventana1 = Toplevel()
        self.ventana1.geometry("500x500+500+100")
        self.ventana1.title("Horas Extra")
        self.ventana1.resizable(0,0)
        self.ventana1.config(bg="blue")
        self.ventana1.iconbitmap("Adds/icon.ico")
        self.ventana1.config(bd=25)
        self.ventana1.config(relief="groove")
        self.ventana1.config(cursor="pirate")
        self.ventana1.config(bd=25)
        self.ventana1.config(relief="groove")
        self.ventana1.config(cursor="pirate")

        self.horas = StringVar()

        self.label2 = Label(self.ventana1, text="Horas", font=("Arial", 15), bg="blue", fg="white")
        self.label2.place(x=50, y=100)
        self.entrada2 = Entry(self.ventana1, textvariable=self.horas, font=("Arial", 15))
        self.entrada2.place(x=150, y=100)

        self.boton1 = Button(self.ventana1, text="Calcular solo uno", font=("Arial", 15), command=self.calcular_horas)
        self.boton1.place(x=200, y=150)

        self.boton2 = Button(self.ventana1, text="Calcular Todo", font=("Arial", 15), 
        command=self.calcular_todo_horas)
        self.boton2.place(x=200, y=200)

        self.boton3 = Button(self.ventana1, text="Calcular Algunos", font=("Arial", 15), 
        command=self.algunos)
        self.boton3.place(x=200, y=250)

    #############################Boton 1#############################
        """
        def calcular_horas(self): Esta funcion se encarga de calcular las horas extra de un empleado en especifico
        E: seleciona un empleado de la lista y se coloca las horas extra que trabajo
        S: se muestra en una tabla las horas extra que trabajo el empleado
        R: solo se puede seleccionar un empleado de la lista
        """
    def calcular_horas(self):
            horas = self.horas.get()
            listbox = Listbox(self.ventana1, width=50, height=10)
            listbox.place(x=50, y=200)
            for i in lista:
                listbox.insert(END, i[2] + " " + i[0]+ " " + i[1])
                #Solo se puede seleccionar un empleado de la lista
                listbox.bind("<<ListboxSelect>>", self.onselect2)
            self.boton4 = Button(self.ventana1, text="Calcular 1", font=("Arial", 15), command=self.calcular_uno)
            self.boton4.place(x=200, y=300)

    """
    onselect2(self, event): Esta funcion se encarga de seleccionar el empleado que se va a calcular las horas extra
    """
    def onselect2(self, event):
        widget = event.widget
        selection = widget.curselection()
        picked = widget.get(selection[0])
        lista2.append(picked[0:4])
        print(lista2)


    """  
    def calcular_uno(self): Esta funcion se encarga de calcular las horas extra de un empleado en especifico, aqui 
    se le suman lo extra que trabajo
    """
    def calcular_uno(self): 
        horas = self.horas.get()
        for i in lista:
            for j in lista2:
                if i[2] == j:
                    horas = float(horas)
                    if horas > 8:
                        messagebox.showinfo("Horas Extra", 
                        "El empleado " + i[0] + " " + i[2] + "No puede llevar mas de 8 horas extra")
                    else:
                        i[7] = float(i[7]) + float(horas)
                        i[7] = str(i[7])
                        messagebox.showinfo("Horas Extra", 
                        "El empleado " + i[0] + " " + i[2] + " tiene " + i[7] + " horas extra")
                        ventana = C_Salario(self.ventana)
                        lista2.clear()
                        print(lista2)
                        print("listo")
                        self.ventana1.destroy()

                


    ############################# Boton 2 #############################
    """
    def calcular_todo_horas(self): Esta funcion se encarga de calcular las horas extra de todos los empleados
    E: se coloca las horas extra que trabajo cada empleado
    S: se muestra en una tabla las horas extra que trabajo cada empleado
    R: solo se puede seleccionar un empleado de la lista
    """
    def calcular_todo_horas(self):
            self.ventana2 = Toplevel()
            self.ventana2.geometry("500x500+500+100")
            self.ventana2.title("Horas Extra Todos")
            self.ventana2.resizable(0,0)
            self.ventana2.config(bg="blue")
            self.ventana2.iconbitmap("Adds/icon.ico")
            self.ventana2.config(bd=25)
            self.ventana2.config(relief="groove")
            self.ventana2.config(cursor="pirate")
            self.ventana2.config(bd=25)
            self.ventana2.config(relief="groove")
            self.ventana2.config(cursor="pirate")

            self.horas = StringVar()

            self.label1 = Label(self.ventana2, text="Horas Extra", font=("Arial", 15), bg="blue", fg="white")
            self.label1.place(x=50, y=50)
            self.entrada1 = Entry(self.ventana2, textvariable=self.horas, font=("Arial", 15))
            self.entrada1.place(x=175, y=50)

            self.boton1 = Button(self.ventana2, text="Calcular", font=("Arial", 15), command=self.calcular_todo_horas2)
            self.boton1.place(x=200, y=100)

            mensaje = StringVar()
            mensaje.set(" ")
            self.label2 = Label(self.ventana2, textvariable=self.mensaje, font=("Arial", 15), bg="blue", fg="white")
            self.label2.place(x=50, y=150)
    
    """
    def calcular_todo_horas2(self): Esta funcion se encarga de calcular el salario de todos los empleados
    """
    def calcular_todo_horas2(self):
        horas = self.horas.get()
        for i in lista:
            self.total = float(i[7]) + float(horas)
            i[7] = self.total
            escribir()
            self.mensaje.set("Suma Hecha Satisfactoriamente")
            self.total = 0

    ############################# Boton 3 #############################

    """
    def algunos(self): Esta funcion se encarga de calcular las horas extra de algunos empleados
     """
    def algunos(self):
            self.ventana3 = Toplevel()
            self.ventana3.geometry("500x500+500+100")
            self.ventana3.title("Horas Extra Algunos")
            self.ventana3.resizable(0,0)
            self.ventana3.config(bg="blue")
            self.ventana3.iconbitmap("Adds/icon.ico")
            self.ventana3.config(bd=25)
            self.ventana3.config(relief="groove")
            self.ventana3.config(cursor="pirate")
            self.ventana3.config(bd=25)
            self.ventana3.config(relief="groove")
            self.ventana3.config(cursor="pirate")

            mensaje= StringVar()
            mensaje.set(" ")

            self.label2 = Label(self.ventana3, text="Horas", font=("Arial", 15), bg="blue", fg="white")
            self.label2.place(x=50, y=100)

            self.entrada2 = Entry(self.ventana3, textvariable=self.horas, font=("Arial", 15))
            self.entrada2.place(x=150, y=100)

            self.boton1 = Button(self.ventana3, text="Calcular", font=("Arial", 15), command=self.calcular_algunos)
            self.boton1.place(x=200, y=150)



    """
    def calcular_algunos(self): Esta funcion se encarga de calcular las horas extra de algunos empleados
    que se van a seleccionar en esta ventana
    """
    def calcular_algunos(self):
            horas = self.horas.get()
            listbox = Listbox(self.ventana3, width=50, height=10)
            listbox.place(x=50, y=200)
            for i in lista:
                listbox.insert(END, i[2] + " " + i[0] + " " + i[1])
            #Se va a seleccionar de manera multiple los empleados que se van a calcular las horas extra
            listbox.config(selectmode=MULTIPLE)
            listbox.bind("<<ListboxSelect>>", self.onselect)
            listbox.pack()
            #resive la informacion de la funcion onselect que es la lista de los empleados seleccionados
            self.boton2 = Button(self.ventana3, text="Calcular", font=("Arial", 15), command=self.calcular_algunos2)
            self.boton2.place(x=200, y=400)

    """
    def onselect(self, evt): Esta funcion se encarga de guardar en una lista los empleados seleccionados
    pero solo la parte del codigo  que son los primeros 4 caracteres, de cada empleado seleccionado, pero solo
    el codigo y una vez nada mas se añade el codigo de dicho empleado a la lista2 para asi poder calcular las horas
    del empleado para asi poder calcular las horas extra de los empleados
    seleccionados en la funcion calcular_algunos2
    """
    def onselect(self, evt):
        w = evt.widget
        index = int(w.curselection()[0])
        value = w.get(index)
        print('You selected item %d: "%s"' % (index, value))
        lista2.append(value[0:4])
        print(lista2)
    
    """
    def calcular_algunos2(self): Esta funcion se encarga de calcular las horas extra de los empleados seleccionados
    esta resive la informacion de la lista2 y de la funcion calcular_algunos para calcular las horas extra
    de los empleados seleccionados despues de seleccionar los empleados se debe dar click en el boton calcular, este 
    debe de dar un mensaje al final cuando haya calculado el de todas las horas extra de los empleados seleccionados
    y de sumarlas a sus horas extra totales
    """
    def calcular_algunos2(self):
        horas = self.horas.get()
        for i in lista:
            for j in lista2:
                if i[2] == j:
                    self.total = float(i[7]) + float(horas)
                    i[7] = self.total
                    escribir()
                    self.mensaje.set("Suma Hecha Satisfactoriamente")
                    self.total = 0
                    ventana = C_Salario(self.ventana)
        lista2.clear()

    ############################# Boton 4 #############################
    def volver(self):
        #self.ventana.destroy()
        self.ventana = Ventana_Jefe(self.ventana)

    ############################# Boton 5 #############################
    """
    def graficas(self): Esta funcion se encarga de mostrar las graficas de los empleados, de las horas extra
    de los empleados, de los salarios de los empleados, de los salarios de los empleados con las horas extra
    y de los salarios de los empleados sin las horas extra tambien del genero de los empleados
    """
    def graficas(self):
        self.ventana4 = Toplevel()
        self.ventana4.geometry("500x500+500+100")
        self.ventana4.title("Graficas")
        self.ventana4.resizable(0,0)
        self.ventana4.config(bg="blue")
        self.ventana4.iconbitmap("Adds/icon.ico")
        self.ventana4.config(bd=25)
        self.ventana4.config(relief="groove")
        self.ventana4.config(cursor="pirate")
        self.ventana4.config(bd=25)
        self.ventana4.config(relief="groove")
        self.ventana4.config(cursor="pirate")

        self.boton1 = Button(self.ventana4, text="Horas Semana", font=("Arial", 15), command=self.grafica1)
        self.boton1.place(x=50, y=100)

        self.boton2 = Button(self.ventana4, text="Edad", font=("Arial", 15), command=self.grafica2)
        self.boton2.place(x=50, y=150)

        self.boton3 = Button(self.ventana4, text="Horas Extra", font=("Arial", 15), command=self.grafica3)
        self.boton3.place(x=50, y=200)

        self.boton4 = Button(self.ventana4, text="Salario Total", font=("Arial", 15), command=self.grafica4)
        self.boton4.place(x=50, y=250)

        self.boton5 = Button(self.ventana4, text="Genero", font=("Arial", 15), command=self.grafica5)
        self.boton5.place(x=50, y=300)

        self.boton6 = Button(self.ventana4, text="Volver", font=("Arial", 15), command=self.volver2)
        self.boton6.place(x=50, y=350)

    ############################# Boton 6  de las graficas#############################
    def volver2(self):
        self.ventana4.destroy()
        self.ventana = C_Salario(self.ventana)

    ############################# Grafica 1 #############################
    """
    def grafica1(self): Esta funcion se encarga de mostrar la grafica de las horas extra de los empleados
    en 5 tipos de horas extra: 0-5, 6-10, 11-15, 16-20, 21-25
    """
    def grafica1(self):
        self.CeroCinco = 0
        self.SeisDiez = 0
        self.OnceQuince = 0
        self.DieciseisVeinte = 0
        self.VeintiunoVeinticinco = 0
        for i in lista:
            if float(i[4]) < 6.0:
                self.CeroCinco += 1
            elif float(i[4]) < 11.0:
                self.SeisDiez += 1
            elif float(i[4]) < 16.0:
                self.OnceQuince += 1
            elif float(i[4]) < 21.0:
                self.DieciseisVeinte += 1
            else:
                self.VeintiunoVeinticinco += 1
        
        self.x = ["0-5", "6-10", "11-15", "16-20", "21-25"]
        self.y = [self.CeroCinco, self.SeisDiez, self.OnceQuince, self.DieciseisVeinte, self.VeintiunoVeinticinco]
        plt.bar(self.x, self.y, color="blue")
        plt.title("Horas ")
        plt.xlabel("Horas ")
        plt.ylabel("Cantidad de Empleados")
        plt.show()

    ############################# Grafica 2 #############################
    """
    def grafica2(self): Esta funcion se encarga de mostrar la grafica en cuanto a la edad de los empleados
    en 5 tipos de edades: 18-25, 26-35, 36-45, 46-55, 56-65
    """
    def grafica2(self):
        self.DieciochoVeinticinco = 0
        self.VeintiseisTreintaCinco = 0
        self.TreintaSeisCuarentaCinco = 0
        self.CuarentaSeisCincuentaCinco = 0
        self.CincuentaSeisSesentaCinco = 0
        for i in lista:
            if int(i[6]) < 26:
                self.DieciochoVeinticinco += 1
            elif int(i[6]) < 36:
                self.VeintiseisTreintaCinco += 1
            elif int(i[6]) < 46:
                self.TreintaSeisCuarentaCinco += 1
            elif int(i[6]) < 56:
                self.CuarentaSeisCincuentaCinco += 1
            elif int(i[6]) < 66:
                self.CincuentaSeisSesentaCinco += 1
        #Logica de la grafica usando matplotlib
        self.x = ["18-25", "26-35", "36-45", "46-55", "56-65"]
        self.y = [self.DieciochoVeinticinco, self.VeintiseisTreintaCinco, 
            self.TreintaSeisCuarentaCinco, self.CuarentaSeisCincuentaCinco, self.CincuentaSeisSesentaCinco]
        plt.bar(self.x, self.y, color="blue")
        plt.title("Grafica de Edad")
        plt.xlabel("Edad")
        plt.ylabel("Cantidad de Empleados")
        plt.show()

    
    ############################# Grafica 3 #############################
    """
    def grafica3(self): Esta funcion se encarga de mostrar la grafica de la cantidad de horas extra de los empleados
    por promedio en 5 tipos en cuanto a horas extra: 0-5, 6-10, 11-15, 16-20, 21-25
    """
    def grafica3(self):
        #Logica de la grafica usando matplotlib
        self.CeroCinco = 0
        self.SeisDiez = 0
        self.OnceQuince = 0
        self.DieciseisVeinte = 0
        self.VeintiunoVeinticinco = 0
        for i in lista:
            if float(i[7]) < 5.0:
                self.CeroCinco += 1
            elif float(i[7]) < 10.0:
                self.SeisDiez += 1
            elif float(i[7]) < 15.0:
                self.OnceQuince += 1
            elif float(i[7]) < 20.0:
                self.DieciseisVeinte += 1
            elif float(i[7]) < 25.0:
                self.VeintiunoVeinticinco += 1

        self.x = ["0-5", "6-10", "11-15", "16-20", "21-25"]
        self.y = [self.CeroCinco, self.SeisDiez, self.OnceQuince, self.DieciseisVeinte, self.VeintiunoVeinticinco]
        plt.bar(self.x, self.y, color="blue")
        plt.title("Grafica de Horas Extra")
        plt.xlabel("Horas Extra")
        plt.ylabel("Empleados")
        plt.show()

    ############################# Grafica 4 #############################
    """
    def grafica4(self): Esta funcion se encarga de mostrar la grafica de los salarios de los empleados 
    en promedio de los empleados, en cuatro tipos: el primero es el salario de los empleados que ganan menos de 1000,
    el segundo es el salario de los empleados que ganan entre 1000 y 2000, el tercero es el salario de los empleados
    que ganan entre 2000 y 3000 y el cuarto es el salario de los empleados que ganan mas de 3000
    """
    def grafica4(self):
        self.menos1000 = 0
        self.entre1000y2000 = 0
        self.entre2000y3000 = 0
        self.mas3000 = 0
        for i in lista:
            if float(i[9]) < 1000:
                self.menos1000 += 1
            elif float(i[9]) >= 1000 and float(i[9]) < 2000:
                self.entre1000y2000 += 1
            elif float(i[9]) >= 2000 and float(i[9]) < 3000:
                self.entre2000y3000 += 1
            elif float(i[9]) >= 3000:
                self.mas3000 += 1

        self.x = ["Menos de 1000", "Entre 1000 y 2000", "Entre 2000 y 3000", "Mas de 3000"]
        self.y = [self.menos1000, self.entre1000y2000, self.entre2000y3000, self.mas3000]
        plt.bar(self.x, self.y, color="blue")
        plt.title("Grafica de Salario sin Horas Extra")
        plt.xlabel("Salario")
        plt.ylabel("Empleados")
        plt.show()

    ############################# Grafica 5 #############################
    """
    def grafica5(self): Esta funcion se encarga de mostrar la grafica de los generos de los empleados
    la grafica sera circular,los datos se mostraran en porcentajes y son strings para eso contara cuantos
    hombres ,otros y mujeres hay y los convertira en porcentajes de cuantos hay de cada uno
    """
    def grafica5(self):
        #Logica de la grafica usando matplotlib
        self.h = 0
        self.m = 0
        self.o = 0
        for i in lista:
            if i[5] == "Masculino":
                self.h += 1
            elif i[5] == "Femenino":
                self.m += 1
            elif i[5] == "Otro":
                self.o += 1
        self.total = self.h + self.m + self.o
        self.porcentajeH = (self.h * 100) / self.total
        self.porcentajeM = (self.m * 100) / self.total
        self.porcentajeO = (self.o * 100) / self.total
        self.porcentajes = [self.porcentajeH, self.porcentajeM, self.porcentajeO]
        self.nombres = ["Hombres", "Mujeres", "Otros"]
        plt.pie(self.porcentajes, labels=self.nombres, autopct="%0.1f %%")
        plt.title("Grafica de Generos")
        plt.show()

#Calcula el salario de un empleado en especifico 
"""
class Salario_Empleado(): Esta clase es la que se encarga de calcular el salario de un empleado en especifico
"""  
class Salario_Empleado():

    def __init__(self,ventana):

        self.Ventana = ventana.withdraw()
        self.ventana = Toplevel()
        self.ventana.geometry("500x500+500+100")
        self.ventana.title("Calcular Salario")
        self.ventana.resizable(0,0)
        self.ventana.config(bg="Black")
        self.ventana.iconbitmap("Adds/icon.ico")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        
        self.codigo = StringVar()
        self.mensaje = StringVar()
        self.reduccion = 0
        self.total = 0

        #Labels, Entrys y Botones

        self.label1 = Label(self.ventana, text="Codigo", font=("Arial", 15), bg="blue", fg="white")
        self.label1.place(x=50, y=50)

        self.entrada1 = Entry(self.ventana, textvariable=self.codigo, font=("Arial", 15))
        self.entrada1.place(x=150, y=50)

        self.boton1 = Button(self.ventana, text="Calcular", font=("Arial", 15), command=self.calcular)
        self.boton1.place(x=200, y=100)

        self.boton2 = Button(self.ventana, text="Calcular Todo", font=("Arial", 15), command=self.liquidacion)
        self.boton2.place(x=175, y=150)

        self.label2 = Label(self.ventana, textvariable=self.mensaje, font=("Arial", 15), bg="blue", fg="white")
        self.label2.place(x=50, y=250)

        self.boton3 = Button(self.ventana, text="Volver", font=("Arial", 15), command=self.volver)
        self.boton3.place(x=0, y=0)

    ######### Boton 1 #########
    """
    def calcular(self): Esta funcion se encarga de calcular el salario de un empleado en especifico
    """	
    def calcular(self):
        codigo = self.codigo.get()
        if codigo == "":
            self.mensaje.set("El codigo no puede estar vacio")
            return False
        else:
            for i in lista:
                if codigo == i[2]:
                    messagebox.showinfo("Salario", "El salario de " + i[0] + " " + i[1] +
                    " es de " +  str(float(i[9])))
                    self.codigo.set("")
                    return True
            self.mensaje.set("El codigo no existe")
            messagebox.showerror("Error", "El codigo no existe")
            self.codigo.set("")
            return False
    
    ######### Boton 2 #########

    """	
    def liquidacion(self): Esta funcion se encarga de calcular la liquidacion de un empleado en especifico,
    dependiendo de la fecha de ingreso del empleado se le sumara un monto extra a su salario
    """
    def liquidacion(self):
        codigo = self.codigo.get()
        for i in lista:
            if codigo == i[2]:
                fecha = i[8]
                fecha = fecha.split("/")
                fecha = [int(i) for i in fecha]
                fecha = datetime.date(fecha[2], fecha[1], fecha[0])
                fecha_actual = datetime.date.today()
                diferencia = fecha_actual - fecha
                diferencia = diferencia.days
                diferencia = diferencia/30

                if diferencia > 9:
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: " + str(float(i[9])+3500))
                    self.codigo.set("")

                elif diferencia > 6:
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: " + str(float(i[9])+2500))
                    self.codigo.set("")

                elif diferencia > 3:
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: " + str(float(i[9])+1500))
                    self.codigo.set("")

                else:
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: " + str(float(i[9])+500))
                    self.codigo.set("")
                return True
    
    ######### Boton 3 #########
    def volver(self):
        #self.ventana.destroy()
        self.ventana2 = Ventana_Empleado(self.ventana)

#Clase Ventana_Ordenar lista de empleados
"""
class Ventana_Ordenar(): Esta clase es la que se encarga de ordenar los empleados por sueldo, edad y nombre	
"""
class Ventana_Ordenar():

    def __init__(self):
        self.ventana = Toplevel()
        self.ventana.geometry("500x500+500+100")
        self.ventana.title("Ordenar")
        self.ventana.resizable(0,0)
        self.ventana.config(bg="Black")
        self.ventana.iconbitmap("Adds/icon.ico")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        
        self.orden = StringVar()
        self.mensaje = StringVar()
        self.mensaje.set("")

        #Labels, Combobox y Botones

        self.label1 = Label(self.ventana, text="Ordenar por", font=("Arial", 15), bg="blue", fg="white")
        self.label1.place(x=50, y=50)

        self.combobox1 = ttk.Combobox(self.ventana, textvariable=self.orden, state="readonly")
        self.combobox1["values"] = ("Nombre", "Apellido","Codigo", "Sueldo","Horas por Semana", "Sexo", "Edad"
        , "Horas Extras", "Fecha de Ingreso", "Salario Total")
        self.combobox1.place(x=200, y=50)

        self.boton1 = Button(self.ventana, text="Ordenar", font=("Arial", 15), command=self.ordenar)
        self.boton1.place(x=200, y=100)

        self.label2 = Label(self.ventana, textvariable=self.mensaje, font=("Arial", 15), bg="blue", fg="white")
        self.label2.place(x=50, y=250)
        
        self.titulo = Label(self.ventana, text="Para ver ta tabla \n presione el siguiente boton",
         font=("Arial", 20), bg="black", fg="white")
        self.titulo.place(x=100, y=300)

        self.boton2 = Button(self.ventana, text="Ver Tabla", font=("Arial", 15), command=self.ver_tabla)
        self.boton2.place(x=200, y=400)

        self.boton3 = Button(self.ventana, text="Volver", font=("Arial", 15), command=self.volver)
        self.boton3.place(x=0, y=0)

    ######### Boton 1 #########

    """
    def ordenar(self): Esta funcion se encarga de ordenar los empleados por sueldo, edad y nombre...
    E: los datos de los empleados
    S: los datos de los empleados ordenados
    R: que los datos sean validos
    """	
    def ordenar(self):
        orden = self.orden.get()

        if orden == "Nombre":
            lista.sort(key=lambda lista: lista[0])
            #se le da la vuelta a la lista para que quede de mayor a menor
            lista.reverse()
            self.mensaje.set("Se ordeno por nombre")

        elif orden == "Apellido":
            lista.sort(key=lambda lista: lista[1])
            #se le da la vuelta a la lista para que quede de mayor a menor
            lista.reverse()
            self.mensaje.set("Se ordeno por apellido")

        elif orden == "Codigo":
            lista.sort(key=lambda lista: lista[2])
            #se le da la vuelta a la lista para que quede de mayor a menor
            lista.reverse()
            self.mensaje.set("Se ordeno por codigo")

        elif orden == "Sueldo":
            lista.sort(key=lambda lista: float(lista[3]))
            #se le da la vuelta a la lista para que quede de mayor a menor
            lista.reverse()
            self.mensaje.set("Se ordeno por sueldo")

        elif orden == "Horas por Semana":
            lista.sort(key=lambda lista: float(lista[4]))
            #se le da la vuelta a la lista para que quede de mayor a menor
            lista.reverse()
            self.mensaje.set("Se ordeno por horas por semana")

        elif orden == "Sexo":
            lista.sort(key=lambda lista: lista[5])
            #se le da la vuelta a la lista para que quede de mayor a menor
            lista.reverse()
            self.mensaje.set("Se ordeno por sexo")

        elif orden == "Edad":
            lista.sort(key=lambda lista: Int(lista[6]))
            self.mensaje.set("Se ordeno por edad")

        elif orden == "Horas Extras":
            lista.sort(key=lambda lista: float(lista[7]))
            #se le da la vuelta a la lista para que quede de mayor a menor
            lista.reverse()
            self.mensaje.set("Se ordeno por horas extras")

        elif orden == "Fecha de Ingreso":
            lista.sort(key=lambda lista: lista[8])
            #se le da la vuelta a la lista para que quede de mayor a menor
            lista.reverse()
            self.mensaje.set("Se ordeno por fecha de ingreso")
        
        elif orden == "Salario Total":
            lista.sort(key=lambda lista: float(lista[9]))
            #se le da la vuelta a la lista para que quede de mayor a menor
            lista.reverse()
            self.mensaje.set("Se ordeno por salario total")

        else:
            self.mensaje.set("Seleccione una opcion")

    ######### Boton 2 #########
    """
    def ver_tabla(self): Esta funcion se encarga de mostrar la tabla con los datos ordenados
    E: los datos de los empleados ordenados
    S: la tabla con los datos ordenados
    R: que los datos sean validos
    """

    def ver_tabla(self):
       self.tabla = C_Salario.calcular_todo(self)
    
    ######### Boton 3 #########
    """
    def volver(self): Esta funcion se encarga de volver a la ventana principal
    """	
    def volver(self):
        #self.ventana.destroy()
        self.ventana = Ventana_Jefe(self.ventana)

#Ventana Retirar Empleado
"""	
class Ventana_Retirar(): Esta clase es la que se encarga de retirar un empleado de la lista y calcular su salario total
"""
class Ventana_Retirar():

    def __init__(self):
        self.ventana = Toplevel()
        self.ventana.geometry("500x500+500+100")
        self.ventana.title("Retirar")
        self.ventana.resizable(0,0)
        self.ventana.config(bg="Black")
        self.ventana.iconbitmap("Adds/icon.ico")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")

        self.mensaje = StringVar()
        self.mensaje.set("")

        self.boton1 = Button(self.ventana, text="Retirar un suario", font=("Arial", 15), command=self.retirar)
        self.boton1.place(x=200, y=100)

        self.label2 = Label(self.ventana, textvariable=self.mensaje, font=("Arial", 15), bg="blue", fg="white")
        self.label2.place(x=50, y=250)

        self.boton2 = Button(self.ventana, text="Tu liquidacion", font=("Arial", 15), command=self.obtener_liquidacion)
        self.boton2.place(x=200, y=400)

        self.boton3 = Button(self.ventana, text="Retirar Varios", font=("Arial", 15), command=self.retirarVarios)
        self.boton3.place(x=200, y=300)

        self.boton4 = Button(self.ventana, text="Volver", font=("Arial", 15), command=self.volver)
        self.boton4.place(x=0, y=0)

    ######### Boton 1 #########

    """
    def retirar(self): Esta funcion se encarga de retirar un empleado de la lista y calcular su salario total, el usuario
    podra seleccionar el empleado que desea retirar por medio de su codigo
    """
    def retirar(self):
            listbox = Listbox(self.ventana, width=50, height=10)
            listbox.place(x=50, y=200)
            for i in lista:
                listbox.insert(END, i[2] + " " + i[0] + " " + i[1])
            listbox.bind("<<ListboxSelect>>", self.onselect2)
            self.boton2 = Button(self.ventana, text="Listo", font=("Arial", 15), command=self.retirar_aux)
            self.boton2.place(x=300, y=300)
    
    """
    def onselect2(self): Esta funcion se encarga de seleccionar el empleado que el usuario desea retirar
    """ 
    def onselect2(self, event):
        widget = event.widget
        selection = widget.curselection()
        picked = widget.get(selection[0])
        lista2.append(picked[0:4])
        print(lista2)

    """
    def retirar(self): Esta funcion se encarga de retirar un empleado de la lista y calcular su salario total
    """
    def retirar_aux(self):
        for i in lista:
            if i[2] == lista2[0]:
                    lista.remove(i)
                    self.mensaje.set("Se retiro el empleado")
                    lista2.clear()
                    print("Listo")
                    print(lista2)
                    break
            else:
                pass
        else:
            self.mensaje.set("No se encontro el empleado")

    ######### Boton 2 #########
    """
    def obtener_codigo(self): Esta funcion se encarga de obtener el codigo del empleado que el usuario desea retirar
    """

    def obtener_liquidacion():
            listbox = Listbox(self.ventana, width=50, height=10)
            listbox.place(x=50, y=200)
            for i in lista:
                listbox.insert(END, i[2] + " " + i[0] + " " + i[1])
            listbox.bind("<<ListboxSelect>>", self.onselect3)
            self.boton2 = Button(self.ventana, text="Tu liquidacion", font=("Arial", 15), 
            command=self.CalculeLiquidacion)
            self.boton2.place(x=300, y=400)
    
    """
    def onselect3(self): Esta funcion se encarga de seleccionar el empleado que el usuario desea retirar
    """

    def onselect3(self, evt):
        w = evt.widget
        index = int(w.curselection()[0])
        value = w.get(index)
        self.codigo.set(value[0:4])

    ######### Boton 3 #########
    """ 
    def retirarVarios(self): Esta funcion se encarga de retirar varios empleados de la lista y calcular su salario total
    """
    def retirarVarios(self):
            listbox = Listbox(self.ventana, width=50, height=10)
            listbox.place(x=50, y=200)
            for i in lista:
                listbox.insert(END, i[2] + " " + i[0] + " " + i[1])
            listbox.bind("<<ListboxSelect>>", self.onselect4)
            self.boton2 = Button(self.ventana, text="Listo", font=("Arial", 15), command=self.retirarVarios_aux)
            self.boton2.place(x=300, y=100)

    def onselect4(self, evt):
        w = evt.widget
        index = int(w.curselection()[0])
        value = w.get(index)
        lista2.append(value[0:4])
        print(lista2)

    """
    def retirarVarios_aux(self): Esta funcion se encarga de retirar varios empleados de la lista y 
    calcular su salario total
    """
    def retirarVarios_aux(self):
        for i in lista:
            for j in lista2:
                if i[2] == j:
                        self.calcular_liquidacion()
                        lista.remove(i)
                        self.mensaje.set("Se retiro el empleado")
                        ventana = self.Ventana_Retirar(self.ventana)
                        lista2.clear()
                
                else:
                    pass
        else:
            self.mensaje.set("No se encontro el empleado")



    """	
    def calcular_liquidacion(self): Esta funcion se encarga de calcular el salario total de un empleado,
    si este ha trabajado mas de 3 meses, estos mese se calculan con la fecha de contratacion del empleado.
    Se le da un bono de 1000 dolares y si ha trabajado mas de 6 meses
    se le da un bono de 2000 dolares y si ha trabajado mas de 9 meses se le da un bono de 3000 dolares, pero si no 
    solo se le da el salario total mas un bono de 500 dolares que es el bono que se le da a todos los empleados
    """
    def calcular_liquidacion(self):
        codigo = self.codigo.get()
        for i in lista:
            if i[2] == codigo:
                fecha = i[7]
                fecha = fecha.split("/")
                fecha = [int(i) for i in fecha]
                fecha = datetime.date(fecha[2], fecha[1], fecha[0])
                fecha_actual = datetime.date.today()
                diferencia = fecha_actual - fecha
                diferencia = diferencia.days
                diferencia = diferencia/30
                if diferencia > 9:
                    i.append(str(float(i[4])*float(i[3])+3500))
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: "+str(i[8]))
                    self.codigo.set("")

                elif diferencia > 6:
                    i.append(str(float(i[4])*float(i[3])+2500))
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: "+str(i[8]))
                    self.codigo.set("")

                elif diferencia > 3:
                    i.append(str(float(i[4])*float(i[3])+1500))
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: "+str(i[8]))
                    self.codigo.set("")

                else:
                    i.append(str(float(i[4])*float(i[3])+500))
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: "+str(i[8]))
                    self.codigo.set("")

    """
    def CalculeLiquidacion(self): Esta funcion se encarga de calcular el salario total de un empleado, si este ha 
    trabajado mas de 3 meses, estos mese se calculan con la fecha de contratacion del empleado. 
    Se le da un bono de 1000 dolares y si ha trabajado mas de 6 meses se le da un bono de 2000 dolares 
    y si ha trabajado mas de 9 meses se le da un bono de 3000 dolares, pero si no solo se le da el salario 
    total mas un bono de 500 dolares que es el bono que se le da a todos los empleados
    """
    def CalculeLiquidacion(self):
        codigo = self.codigo.get()
        for i in lista:
            if i[2] == codigo:
                fecha = i[7]
                fecha = fecha.split("/")
                fecha = [int(i) for i in fecha]
                fecha = datetime.date(fecha[2], fecha[1], fecha[0])
                fecha_actual = datetime.date.today()
                diferencia = fecha_actual - fecha
                diferencia = diferencia.days
                diferencia = diferencia/30
                if diferencia > 9:
                    i.append(str(float(i[4])*float(i[3])+3500))
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: "+str(i[8]))
                    self.codigo.set("")

                elif diferencia > 6:
                    i.append(str(float(i[4])*float(i[3])+2500))
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: "+str(i[8]))
                    self.codigo.set("")

                elif diferencia > 3:
                    i.append(str(float(i[4])*float(i[3])+1500))
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: "+str(i[8]))
                    self.codigo.set("")

                else:
                    i.append(str(float(i[4])*float(i[3])+500))
                    self.mensaje.set("Se calculo el salario total del empleado")
                    messagebox.showinfo("Liquidacion", "El salario total del empleado es: "+str(i[8]))
                    self.codigo.set("")

    #########Boton 4#########
    """
    volver(self): Esta funcion se encarga de volver a la ventana anterior
    """
    def volver(self):
        self.ventana.destroy()
        ventana = self.Ventana_Jefe(self.ventana)

#Clase de la ventana principal
"""
class Ventana_Pricipal(): Esta clase es la que se encarga de crear la ventana principal deonde hbra dos botones
uno para ingresar como empleado y otro para ingresar como administrador o Jefe	
"""
class Ventana_Pricipal():
    def __init__(self, ventana,contador):

        self.ventana = ventana
        self.contador = contador
        self.ventana.title("Menu principal")
        self.ventana.iconbitmap("Adds/icon.ico")
        self.ventana.geometry("500x500+500+100")
        self.ventana.config(bg="black")
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.boton1 = Button(self.ventana, text="Ingresar como empleado", font=("Arial", 15),
         command=self.ingresar_empleado)
        self.boton1.place(x=120, y=30)
        self.boton2 = Button(self.ventana, text="Ingresar como administrador", font=("Arial", 15),
         command=self.Contraseña)
        self.boton2.place(x=110, y=100)
        """	
        self.Logo1 = Imagenes("Logo 1.png",(320,280))
        self.IMG = Label(self.ventana,image = self.Logo1)
        self.IMG.pack()
        self.IMG.place(x = 75, y = 150)
        """	

        self.contador = 0

        self.ventana.mainloop()

    def Contraseña(self):
        """
        Cerrar la ventana anterior y dejar esta con principal
        """
        self.ventana.destroy()
        self.ventana = Tk()
        self.ventana.update()
        self.ventana.iconbitmap("Adds/icon.ico")
        self.ventana.title("Contraseña")
        self.ventana.geometry("500x200+500+100")
        self.ventana.config(bg="Black")
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        self.ventana.config(bd=25)
        self.ventana.config(relief="groove")
        self.ventana.config(cursor="pirate")
        """
        self.Logo2 = Imagenes("One Piece.png",(320,280))
        self.IMG1 = Label(self.ventana,image = self.Logo2)
        self.IMG1.pack()
        self.IMG1.place(x = 75, y = 150)
        """	
        self.label1 = Label(self.ventana, text="Contraseña", font=("Arial", 15), bg="salmon4", fg="white")
        self.label1.place(x=50, y=50)

        self.contraseña = StringVar()

        #Hacer que el usuario no pueda ver la contraseña que esta escribiendo

        self.entrada1 = Entry(self.ventana, textvariable=self.contraseña, font=("Arial", 15), show="*")
        self.entrada1.place(x=200, y=50)

        self.boton1 = Button(self.ventana, text="Ingresar", font=("Arial", 15), command=self.ingresar_administrador)
        self.boton1.place(x=250, y=100)

        self.boton2 = Button(self.ventana, text="Volver", font=("Arial", 15), command=self.Volver)
        self.boton2.place(x=0, y=0)
    
    """
    def ingresar_administrador(self): Esta funcion se encarga de verificar si la contraseña es correcta y si lo es
    llama a la clase Ventana_Jefe
    """
    def ingresar_administrador(self):

        self.ventana.update()

        contraseña = self.contraseña.get()
        
        if contraseña == "One Piece":
            self.contador = 0
            contraseña = self.contraseña.set(" ")
            self.llamar()

        else:
            messagebox.showerror("Error", "Contraseña incorrecta")
            contraseña = self.contraseña.set(" ")
            self.contador += 1

            if self.contador == 3:
                messagebox.showerror("Error", "Ha ingresado mal la contraseña 3 veces, el programa se cerrara")
                contraseña = self.contraseña.set(" ")
                self.ventana.destroy()

            else:
                pass
    
    """	
    def llamar(self): Esta funcion se encarga de destruir la ventana de contraseña y llamar a la clase Ventana_Jefe
    """	
    def llamar(self):
        self.ventana2 = Ventana_Jefe(self.ventana)

    """
    def ingresar_empleado(self): Esta funcion se encarga de llamar a la clase Ventana_Empleado
    """
    def ingresar_empleado(self):
        self.ventana2 = Ventana_Empleado(self.ventana)
    
    """
    def Volver(self): Esta funcion se encarga de destruir la ventana de contraseña y volver a la ventana principal
    """
    def Volver(self):
        self.ventana.destroy()
        self.Ventana = Aplicacion()

#Clase de ejecucion
"""
class Aplicacion(): Esta clase es la que se encarga de crear la ventana y llamar a la clase Ventana
"""
class Aplicacion():
    def __init__(self):
        self.ventana = Tk()
        self.ventana_principal = Ventana_Pricipal(self.ventana,0)
        self.ventana.mainloop()

#Funciones y clases para el funcionamiento del programa

leer()
leer2()
Salario_total()
Resetear_H_E()
actualizar()
Excel()
aplicacion = Aplicacion()
